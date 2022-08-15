"""Box for choosing from which file to load the input data."""
from pathlib import Path
from typing import Callable, Optional

import toga
from openpyxl import load_workbook
from toga.style import Pack

from eddington_gui.boxes.line_box import LineBox
from eddington_gui.consts import NO_VALUE, SMALL_PADDING


class InputFileBox(LineBox):  # pylint: disable=too-many-instance-attributes
    """Visual box instance for choosing input file."""

    __input_file_path: toga.TextInput
    __select_file_button: toga.Button
    __sheet_label: toga.Label
    __sheet_selection: toga.Selection

    __sheet_selection_enabled: bool
    __on_input_file_change: Optional[Callable[[], None]]

    on_csv_read: Optional[Callable]
    on_excel_read: Optional[Callable]
    on_select_excel_file: Optional[Callable]

    def __init__(self, on_choose_record):
        """Initialize box."""
        super().__init__()
        self.__sheet_selection_enabled = False
        self.on_input_file_change = None
        self.on_csv_read = None
        self.on_excel_read = None
        self.on_select_excel_file = None

        self.__input_file_path = toga.TextInput(readonly=True, style=Pack(flex=1))
        self.__select_file_button = toga.Button(
            label="Choose file",
            on_press=self.select_file,
            style=Pack(padding_left=SMALL_PADDING),
        )
        self.add(
            toga.Label(text="Input file:"),
            self.__input_file_path,
            self.__select_file_button,
            toga.Button(
                label="Choose records",
                on_press=on_choose_record,
                style=Pack(padding_left=SMALL_PADDING),
            ),
        )

        self.__sheet_label = toga.Label(text="Sheet:")
        self.__sheet_selection = toga.Selection(on_select=self.select_sheet)

    @property
    def file_path(self):
        """Getter for the chosen file path."""
        return self.__input_file_path.value

    @file_path.setter
    def file_path(self, file_path):
        """
        Setter for the chosen file path.

        Once a path has been chosen, run handlers to notify other components of the
        change.
        """
        if file_path is None:
            self.__input_file_path.value = ""
        else:
            self.__input_file_path.value = str(file_path)
        if self.on_input_file_change is not None:
            self.on_input_file_change()

    @property
    def sheets_options(self):
        """Sheets options getter. Relevant for excel files."""
        return self.__sheet_selection.items

    @sheets_options.setter
    def sheets_options(self, options):
        """Sheets options setter. Relevant for excel files."""
        if options is None:
            self.__sheet_selection.items = []
            self.sheet_selection_enabled = False

        else:
            self.__sheet_selection.items = options
            self.sheet_selection_enabled = True

    @property
    def sheet_selection_enabled(self):
        """Whether sheet selection is enabled or not."""
        return self.__sheet_selection_enabled

    @sheet_selection_enabled.setter
    def sheet_selection_enabled(self, sheet_selection_enabled):
        """Set/unset sheet selection."""
        old_enabled = self.__sheet_selection_enabled
        self.__sheet_selection_enabled = sheet_selection_enabled
        if old_enabled and not sheet_selection_enabled:
            self.remove(self.__sheet_label, self.__sheet_selection)
        if not old_enabled and sheet_selection_enabled:
            self.insert(2, self.__sheet_label)
            self.insert(3, self.__sheet_selection)

    @property
    def on_input_file_change(self) -> Optional[Callable]:
        """on_input_file_change getter."""
        return self.__on_input_file_change

    @on_input_file_change.setter
    def on_input_file_change(self, on_input_data_change):
        """on_input_file_change setter."""
        self.__on_input_file_change = on_input_data_change

    def select_file(self, widget):  # pylint: disable=unused-argument
        """Open file selection dialog."""
        try:
            input_file_path = Path(
                self.window.open_file_dialog(
                    title="Choose input file", multiselect=False
                )
            )
        except ValueError:
            return
        self.file_path = input_file_path
        suffix = input_file_path.suffix
        if suffix in [".xlsx", ".xls"]:
            excel_file = load_workbook(input_file_path)
            self.sheets_options = [NO_VALUE] + excel_file.sheetnames
            if self.on_select_excel_file is not None:
                self.on_select_excel_file()
            return
        self.sheets_options = None
        if suffix == ".csv":
            if self.on_csv_read is not None:
                self.on_csv_read(input_file_path)
            return
        self.file_path = None
        self.window.error_dialog(
            title="Invalid Input Source",
            message=f"Cannot process file with suffix {suffix}",
        )

    @property
    def selected_sheet(self):
        """Getter for the chosen sheet."""
        return self.__sheet_selection.value

    @selected_sheet.setter
    def selected_sheet(self, selected_sheet):
        """Setter for the chosen sheet."""
        if selected_sheet is None:
            self.__sheet_selection.value = NO_VALUE
        else:
            self.__sheet_selection.value = selected_sheet

    def select_sheet(self, widget):
        """Select sheet to read data from. Relevant for excel files."""
        value = widget.value
        if value == NO_VALUE:
            return
        file_path_value = Path(self.file_path)
        if self.on_excel_read is not None:
            self.on_excel_read(file_path_value, value)
